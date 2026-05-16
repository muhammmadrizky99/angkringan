"""
Simulasi Pelatihan & Evaluasi Model XGBoost - Angkringan Agoy
Mirip persis seperti yang akan dijalankan di Google Colab
"""
import pandas as pd
import numpy as np
from xgboost import XGBRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error
# Time series split: 80% data awal untuk training, 20% data akhir untuk testing
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import warnings
import matplotlib.pyplot as plt
import seaborn as sns
warnings.filterwarnings('ignore')

# ============================================
# 1. LOAD DATA
# ============================================
print("=" * 70)
print("SIMULASI PELATIHAN MODEL XGBOOST - ANGKRINGAN AGOY")
print("Periode Data: 1 Februari 2025 - 30 April 2026 (454 hari)")
print("=" * 70)

df_train_sales = pd.read_excel(r'd:\angkringan\backend\data\data_penjualan_angkringan_setahun.xlsx')
df_train_cuaca = pd.read_csv(r'd:\angkringan\backend\data\data_cuaca_harian.csv')

df_test_sales = pd.read_excel(r'd:\angkringan\backend\data\data_penjualan_april_2026.xlsx')
df_test_cuaca = pd.read_csv(r'd:\angkringan\backend\data\data_cuaca_april_2026.csv')

df_sales = pd.concat([df_train_sales, df_test_sales], ignore_index=True)
df_cuaca = pd.concat([df_train_cuaca, df_test_cuaca], ignore_index=True)

df_sales['tanggal'] = pd.to_datetime(df_sales['tanggal'])
df_cuaca['date'] = pd.to_datetime(df_cuaca['date'])

products = [c for c in df_sales.columns if c != 'tanggal']
print(f"\nJumlah produk: {len(products)}")
print(f"Jumlah total hari (Train + Test): {len(df_sales)}")
print(f"Produk: {', '.join(products)}")

# ============================================
# 2. MERGE & FEATURE ENGINEERING
# ============================================
RAMADAN_1446_START = pd.Timestamp("2025-02-28")
RAMADAN_1446_END = pd.Timestamp("2025-03-29")
RAMADAN_1447_START = pd.Timestamp("2026-02-19")
RAMADAN_1447_END = pd.Timestamp("2026-03-19")

FEATURE_COLS = [
    "day_of_week", "is_weekend", "month", "day_of_month", "is_ramadan",
    "lag_1", "lag_3", "lag_7",
    "rolling_mean_7", "rolling_mean_14", "rolling_std_7",
    "weather", "event"
]

def create_features(df):
    df = df.sort_values("date").copy()
    df["day_of_week"] = df["date"].dt.dayofweek
    df["is_weekend"] = (df["day_of_week"] >= 5).astype(int)
    df["month"] = df["date"].dt.month
    df["day_of_month"] = df["date"].dt.day
    df["is_ramadan"] = (
        ((df["date"] >= RAMADAN_1446_START) & (df["date"] <= RAMADAN_1446_END)) |
        ((df["date"] >= RAMADAN_1447_START) & (df["date"] <= RAMADAN_1447_END))
    ).astype(int)

    # Lag features
    df["lag_1"] = df["quantity"].shift(1)
    df["lag_3"] = df["quantity"].shift(3)
    df["lag_7"] = df["quantity"].shift(7)

    # Rolling statistics
    df["rolling_mean_7"] = df["quantity"].rolling(window=7).mean()
    df["rolling_mean_14"] = df["quantity"].rolling(window=14).mean()
    df["rolling_std_7"] = df["quantity"].rolling(window=7).std()

    df["weather"] = df["weather"].fillna(0).astype(int)
    df["event"] = df["event"].fillna(0).astype(int)
    df = df.dropna(subset=["lag_1", "lag_3", "lag_7", "rolling_mean_7", "rolling_mean_14", "rolling_std_7"])
    return df

def calculate_mape(y_true, y_pred):
    y_true, y_pred = np.array(y_true, dtype=float), np.array(y_pred, dtype=float)
    mask = y_true > 0
    if mask.sum() == 0:
        return 0.0
    return float(np.mean(np.abs((y_true[mask] - y_pred[mask]) / y_true[mask])) * 100)

# ============================================
# 3. TRAIN & EVALUATE PER PRODUK
# ============================================
print("\n" + "=" * 70)
print("HASIL PELATIHAN & EVALUASI MODEL")
print("=" * 70)
print(f"{'Produk':<28} {'MAE':>6} {'RMSE':>7} {'MAPE(%)':>8} {'Akurasi':>8} {'Status'}")
print("-" * 75)

all_results = []
all_predictions = []

for product in products:
    # Siapkan data per produk
    df_prod = pd.DataFrame({
        'date': df_sales['tanggal'],
        'quantity': df_sales[product]
    })
    df_prod = df_prod.merge(df_cuaca, left_on='date', right_on='date', how='left')
    
    # Hapus hari tutup (quantity = 0 untuk semua produk)
    # Tapi tetap simpan untuk konteks lag
    df_features = create_features(df_prod)
    
    if len(df_features) < 20:
        print(f"  {product:<28} {'SKIP - data terlalu sedikit'}")
        continue
    
    # SPLIT TRAINING (Feb 2025 - Mar 2026) vs TESTING (April 2026)
    train_mask = df_features['date'] < '2026-04-01'
    test_mask = df_features['date'] >= '2026-04-01'
    
    X_train = df_features.loc[train_mask, FEATURE_COLS]
    y_train = df_features.loc[train_mask, "quantity"]
    
    X_test = df_features.loc[test_mask, FEATURE_COLS]
    y_test = df_features.loc[test_mask, "quantity"]
    
    if len(X_test) == 0:
        continue
    
    # Train model
    model = XGBRegressor(
        n_estimators=200,
        max_depth=5,
        learning_rate=0.05,
        subsample=0.85,
        colsample_bytree=0.85,
        min_child_weight=2,
        gamma=0.05,
        reg_alpha=0.05,
        reg_lambda=0.8,
        random_state=42,
        objective="reg:squarederror",
    )
    model.fit(X_train, y_train, verbose=False)
    
    # Predict untuk April 2026
    y_pred = model.predict(X_test)
    y_pred = np.maximum(y_pred, 0)
    
    # Metrics Evaluasi (dibandingkan dengan data Aktual April)
    mae = mean_absolute_error(y_test, y_pred)
    rmse = np.sqrt(mean_squared_error(y_test, y_pred))
    mape = calculate_mape(y_test.values, y_pred)
    accuracy = 100 - mape
    
    # Interpretasi
    if mape < 15:
        status = "Sangat Baik"
    elif mape < 25:
        status = "Baik"
    elif mape < 40:
        status = "Cukup"
    else:
        status = "Perlu Improve"
        
    # Simpan hasil untuk diexport ke Excel (hanya hari buka, aktual > 0)
    for idx, (date, actual, pred) in enumerate(zip(df_features.loc[test_mask, 'date'], y_test, y_pred)):
        if int(actual) > 0:  # Abaikan hari tutup agar MAPE tidak inkonsisten
            all_predictions.append({
                'Tanggal': date.strftime('%Y-%m-%d'),
                'Produk': product,
                'Data Asli (Aktual)': int(actual),
                'Prediksi XGBoost': round(pred)
            })
    
    all_results.append({
        'product': product,
        'mae': mae,
        'rmse': rmse,
        'mape': mape,
        'accuracy': accuracy,
        'status': status,
        'importance': dict(zip(FEATURE_COLS, [float(x) for x in model.feature_importances_])),
        'y_test': y_test.values,
        'y_pred': y_pred
    })
    
    print(f"  {product:28} {mae:6.2f} {rmse:7.2f} {mape:7.2f}% {accuracy:7.2f}% {status}")

# ============================================
# 4. RINGKASAN KESELURUHAN
# ============================================
print("\n" + "=" * 70)
print("RINGKASAN METRIK KESELURUHAN")
print("=" * 70)

avg_mae = np.mean([r['mae'] for r in all_results])
avg_rmse = np.mean([r['rmse'] for r in all_results])
avg_mape = np.mean([r['mape'] for r in all_results])

print(f"  Rata-rata MAE      : {avg_mae:.2f} porsi")
print(f"  Rata-rata RMSE     : {avg_rmse:.2f} porsi")
print(f"  Rata-rata MAPE     : {avg_mape:.2f}%")
print(f"  Rata-rata Akurasi  : {100 - avg_mape:.2f}%")

sangat_baik = sum(1 for r in all_results if r['mape'] < 15)
baik = sum(1 for r in all_results if 15 <= r['mape'] < 25)
cukup = sum(1 for r in all_results if 25 <= r['mape'] < 40)
kurang = sum(1 for r in all_results if r['mape'] >= 40)

print(f"\n  Distribusi Kualitas:")
print(f"    Sangat Baik (MAPE < 15%)  : {sangat_baik}/{len(all_results)} produk")
print(f"    Baik (MAPE 15-25%)        : {baik}/{len(all_results)} produk")
print(f"    Cukup (MAPE 25-40%)       : {cukup}/{len(all_results)} produk")
print(f"    Perlu Improve (MAPE > 40%): {kurang}/{len(all_results)} produk")

# ============================================
# 5. FEATURE IMPORTANCE (rata-rata semua produk)
# ============================================
print("\n" + "=" * 70)
print("FEATURE IMPORTANCE (RATA-RATA SEMUA PRODUK)")
print("=" * 70)

avg_importance = {}
for feat in FEATURE_COLS:
    avg_importance[feat] = np.mean([r['importance'][feat] for r in all_results])

sorted_imp = sorted(avg_importance.items(), key=lambda x: x[1], reverse=True)
for feat, imp in sorted_imp:
    bar = "#" * int(imp * 100)
    print(f"  {feat:<20} {imp:.4f} {bar}")

# ============================================
# 6. CONTOH PREDIKSI vs AKTUAL (3 produk utama)
# ============================================
print("\n" + "=" * 70)
print("CONTOH PREDIKSI vs AKTUAL (5 sample dari test set)")
print("=" * 70)

for product in ['bakso', 'ceker', 'telur_puyuh']:
    df_prod = pd.DataFrame({
        'date': df_sales['tanggal'],
        'quantity': df_sales[product]
    })
    df_prod = df_prod.merge(df_cuaca, left_on='date', right_on='date', how='left')
    df_features = create_features(df_prod)
    
    # SPLIT TRAINING (Feb 2025 - Mar 2026) vs TESTING (April 2026)
    train_mask = df_features['date'] < '2026-04-01'
    test_mask = df_features['date'] >= '2026-04-01'
    
    X_train = df_features.loc[train_mask, FEATURE_COLS]
    y_train = df_features.loc[train_mask, "quantity"]
    
    X_test = df_features.loc[test_mask, FEATURE_COLS]
    y_test = df_features.loc[test_mask, "quantity"]
    
    model = XGBRegressor(
        n_estimators=200, max_depth=5, learning_rate=0.05,
        subsample=0.85, colsample_bytree=0.85, min_child_weight=2,
        gamma=0.05, reg_alpha=0.05, reg_lambda=0.8,
        random_state=42, objective="reg:squarederror",
    )
    model.fit(X_train, y_train, verbose=False)
    y_pred = np.maximum(model.predict(X_test), 0)
    
    print(f"\n  [{product.upper()}]")
    print(f"  {'No':>4} {'Aktual':>7} {'Prediksi':>9} {'Selisih':>8}")
    
    for i in range(min(5, len(y_test))):
        actual = int(y_test.iloc[i])
        predicted = round(y_pred[i])
        diff = predicted - actual
        print(f"  {i+1:>4} {actual:>7} {predicted:>9} {diff:>+8}")

print("\n" + "=" * 70)
print("SIMULASI SELESAI - SIAP UNTUK GOOGLE COLAB")
print("=" * 70)

# ============================================
# 7. GENERATE EXCEL EVALUASI
# ============================================
if len(all_predictions) > 0:
    print("\nMenghasilkan file Excel Evaluasi...")
    df_results = pd.DataFrame(all_predictions)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluasi April 2026"

    headers = ["Tanggal", "Produk", "Data Asli (Aktual)", "Prediksi XGBoost", "Selisih (MAE)", "Persentase Error (MAPE)"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(df_results.itertuples(index=False), 2):
        ws.cell(row=r_idx, column=1, value=row[0])
        ws.cell(row=r_idx, column=2, value=row[1])
        ws.cell(row=r_idx, column=3, value=row[2])
        ws.cell(row=r_idx, column=4, value=row[3])
        ws.cell(row=r_idx, column=5, value=f"=ABS(C{r_idx}-D{r_idx})")
        ws.cell(row=r_idx, column=6, value=f"=IF(C{r_idx}>0, ABS(C{r_idx}-D{r_idx})/C{r_idx}, 0)")
        ws.cell(row=r_idx, column=6).number_format = '0.00%'

    last_row = len(df_results) + 1
    ws.cell(row=last_row + 2, column=5, value="Rata-rata MAPE Keseluruhan:")
    ws.cell(row=last_row + 2, column=5).font = Font(bold=True)
    ws.cell(row=last_row + 2, column=6, value=f"=AVERAGE(F2:F{last_row})")
    ws.cell(row=last_row + 2, column=6).font = Font(bold=True)
    ws.cell(row=last_row + 2, column=6).number_format = '0.00%'
    ws.cell(row=last_row + 2, column=6).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25

    excel_output = r'd:\angkringan\tugasakhir\Evaluasi_Prediksi_April_2026.xlsx'
    wb.save(excel_output)
    print(f"✅ Excel Evaluasi berhasil dibuat di: {excel_output}")

    # ============================================
    # 8. VISUALISASI (CHARTS) - Tampil di Output Colab
    # ============================================
    print("\nMenampilkan grafik visualisasi...")
    
    # A. Grafik MAPE & Akurasi per Produk
    plt.figure(figsize=(12, 6))
    product_names = [r['product'] for r in all_results]
    mape_values = [r['mape'] for r in all_results]
    colors = ['#2ecc71' if m < 15 else '#f39c12' if m < 25 else '#e74c3c' for m in mape_values]
    bars = plt.barh(product_names, mape_values, color=colors)
    plt.xlabel('MAPE (%)')
    plt.title('Perbandingan MAPE per Produk')
    plt.axvline(x=15, color='gray', linestyle='--', alpha=0.5, label='Batas Sangat Baik (15%)')
    plt.legend()
    plt.tight_layout()
    plt.show()

    # B. Feature Importance Chart
    plt.figure(figsize=(10, 6))
    feat_importances = pd.DataFrame([res['importance'] for res in all_results]).mean().sort_values(ascending=False)
    sns.barplot(x=feat_importances.values, y=feat_importances.index, palette='viridis')
    plt.title('Faktor-Faktor yang Mempengaruhi Prediksi (Feature Importance)')
    plt.xlabel('Tingkat Pengaruh (Score)')
    plt.tight_layout()
    plt.show()

    # C. Prediksi vs Aktual (3 Produk Utama)
    sample_products = ['bakso', 'ceker', 'telur_puyuh']
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    for ax, prod_name in zip(axes, sample_products):
        res = next((r for r in all_results if r['product'] == prod_name), None)
        if res:
            ax.plot(res['y_test'], label='Aktual', marker='o', linewidth=2)
            ax.plot(res['y_pred'], label='Prediksi', marker='s', linestyle='--', color='orange')
            ax.set_title(f'{prod_name.upper()} (MAPE: {res["mape"]:.1f}%)')
            ax.set_xlabel('Hari ke-')
            ax.set_ylabel('Jumlah Porsi')
            ax.legend()
            ax.grid(True, alpha=0.3)
    plt.suptitle('Grafik Perbandingan Prediksi vs Aktual - April 2026', fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.show()
    
    print("\n" + "=" * 70)
    print("SELURUH PROSES SELESAI!")
    print("=" * 70)
